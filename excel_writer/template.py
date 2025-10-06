from datetime import datetime
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelWriter:
    @property
    def col_length_offset(self) -> int:
        return 12

    @property
    def output_schema(self) -> List[str]:
        raise NotImplementedError("Subclasses should implement this method")

    @property
    def date_columns(self) -> List[str]:
        raise NotImplementedError("Subclasses should implement this method")

    @property
    def number_columns(self) -> List[str]:
        raise NotImplementedError("Subclasses should implement this method")

    @property
    def style_config(self) -> Dict:
        return {
            "header_color": "F2DCDC",
            "header_font_name": "Arial",
            "header_font_size": 12,
            "header_bold": True,
            "data_font_name": "Arial",
            "data_font_size": 12,
        }

    def write_excel(
        self,
        df: pd.DataFrame,
    ):
        temp_file = "temp.xlsx"
        df = df.loc[:, self.output_schema]
        df.to_excel(temp_file, index=False, engine="openpyxl")

        wb = load_workbook(temp_file)
        ws = wb.active

        # Header style
        header_fill = PatternFill(
            start_color=self.style_config["header_color"],
            end_color=self.style_config["header_color"],
            fill_type="solid",
        )
        header_font = Font(
            name=self.style_config["header_font_name"],
            size=self.style_config["header_font_size"],
            bold=self.style_config["header_bold"],
        )

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font

        # Data style
        data_font = Font(
            name=self.style_config["data_font_name"],
            size=self.style_config["data_font_size"],
        )
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.font = data_font

        # Set column format
        for row in range(2, ws.max_row + 1):
            # Data columns
            for col_name in self.date_columns:
                col_idx = df.columns.get_loc(col_name) + 1
                cell = ws.cell(row=row, column=col_idx)
                if isinstance(cell.value, datetime):
                    cell.number_format = "MM-DD-YYYY"

            # Number columns
            for col_name in self.number_columns:
                col_idx = df.columns.get_loc(col_name) + 1
                cell = ws.cell(row=row, column=col_idx)
                if col_name == "Qty":
                    if isinstance(cell.value, str):
                        cell.value = int(cell.value)
                    cell.number_format = "0"
                else:  # Unit Price or others
                    if isinstance(cell.value, str):
                        cell.value = float(cell.value)
                    cell.number_format = "0.00"

        # adjust column width
        for col_idx, col_name in enumerate(df.columns, 1):
            max_length = (
                max(df[col_name].astype(str).map(len).max(), len(col_name))
                + self.col_length_offset
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length

        wb.save(temp_file)

        with open(temp_file, "rb") as f:
            return f.read()
