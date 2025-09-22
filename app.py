import re
from datetime import datetime, timedelta

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

GT_CRD_DAYS = 70
COL_LENGTH_OFFSET = 12


def parse_po_text(text: str, file_type: str = "original") -> dict:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    PO_ID_POSITION = 2
    CREATE_DATE_POSITION = 18
    INFO_POSITION = 25
    if file_type == "revised":
        CREATE_DATE_POSITION += 1
        target_pattern = "Specific changes are in red.."
        for i, line in enumerate(lines):
            if target_pattern in line:
                INFO_POSITION = i + 3
                break

    result = {}

    # PO# ("Purchase Order xxxxxx")
    if len(lines) >= PO_ID_POSITION + 1:
        po_match = re.search(r"Purchase Order\s+([A-Z0-9]+)", lines[PO_ID_POSITION])
        if po_match:
            result["PO#"] = po_match.group(1)

    # Material, Description, Qty, Unit Price (use "EACH" as anchor)
    if len(lines) >= INFO_POSITION + 2:
        mdqu_match = re.search(
            r"\d+\s+\w+\s+([A-Z0-9\-]+)\s+(.+?)\s+(\d+)\s+EACH\s+(\d+\.\d+)", lines[INFO_POSITION]
        )
        if mdqu_match:
            result["Material"] = mdqu_match.group(1)
            desc_part1 = mdqu_match.group(2).strip()
            desc_part2 = lines[INFO_POSITION + 1]
            result["Description"] = (desc_part1 + " " + desc_part2).strip()
            result["Qty"] = mdqu_match.group(3)
            result["Unit Price"] = mdqu_match.group(4)

    # Create Date (first item in the next line of "Date Terms Ship Via")
    if len(lines) >= CREATE_DATE_POSITION + 1:
        date_match = re.search(r"(\d{2}/\d{2}/\d{4})", lines[CREATE_DATE_POSITION])
        if date_match:
            create_date_str = date_match.group(1)
            result["Create Date"] = datetime.strptime(create_date_str, "%m/%d/%Y")

    # Due Date (right after "Delivery Requested Date")
    due_date_match = re.search(r"Delivery Requested Date\s+(\d{2}/\d{2}/\d{4})", text)
    if due_date_match:
        due_date_str = due_date_match.group(1)
        result["Due Date"] = datetime.strptime(due_date_str, "%m/%d/%Y")

    # calculate GT CRD
    if "Create Date" in result.keys():
        result["GT CRD"] = result["Create Date"] + timedelta(days=GT_CRD_DAYS)

    return result


def write_styled_excel(df: pd.DataFrame):
    temp_file = "temp.xlsx"
    df.to_excel(temp_file, index=False, engine="openpyxl")

    wb = load_workbook(temp_file)
    ws = wb.active

    # Header style
    header_fill = PatternFill(start_color="F2DCDC", end_color="F2DCDC", fill_type="solid")
    header_font = Font(name="Arial", size=12, bold=True)

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    # Data style
    data_font = Font(name="Arial", size=12)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = data_font

    # Set column format
    for row in range(2, ws.max_row + 1):
        # Data columns
        for col_name in ["Create Date", "Due Date", "GT CRD"]:
            col_idx = df.columns.get_loc(col_name) + 1
            cell = ws.cell(row=row, column=col_idx)
            if isinstance(cell.value, datetime):
                cell.number_format = "MM-DD-YYYY"

        # Number columns
        if "Qty" in df.columns:
            qty_idx = df.columns.get_loc("Qty") + 1
            cell = ws.cell(row=row, column=qty_idx)
            if isinstance(cell.value, str):
                cell.value = int(cell.value)
            cell.number_format = "0"

        if "Unit Price" in df.columns:
            unit_price_idx = df.columns.get_loc("Unit Price") + 1
            cell = ws.cell(row=row, column=unit_price_idx)
            if isinstance(cell.value, str):
                cell.value = float(cell.value)
            cell.number_format = "0.00"

    # adjust column width
    for col_idx, col_name in enumerate(df.columns, 1):
        max_length = max(df[col_name].astype(str).map(len).max(), len(col_name)) + COL_LENGTH_OFFSET
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length

    wb.save(temp_file)

    with open(temp_file, "rb") as f:
        return f.read()


# -------------------- Streamlit App --------------------
st.set_page_config(layout="centered")  # default

st.markdown(
    """
    <style>
        /* block-container */
        .block-container {
            max-width: 1000px;
            padding-left: 2rem;
            padding-right: 2rem;
        }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("Hi Angel! Welcome to Your Workspace!")
st.subheader("Purchase Order PDF Parser → Excel")

# Initialize session_state
if "df" not in st.session_state:
    st.session_state.df = None
if "excel_bytes" not in st.session_state:
    st.session_state.excel_bytes = None
if "failed_files" not in st.session_state:
    st.session_state.failed_files = []

uploaded_files = st.file_uploader("Please Upload PDF files here", type="pdf", accept_multiple_files=True)
st.write(f"Uploaded files: {len(uploaded_files) if uploaded_files else 0}")

if st.button("Run!"):
    if uploaded_files:
        required_keys = [
            "PO#",
            "Material",
            "Description",
            "Qty",
            "Unit Price",
            "Create Date",
            "Due Date",
            "GT CRD",
        ]
        result_list = []
        revised_result_list = []
        original_files = []
        revised_files = []
        failed_files = []

        for upload_file in uploaded_files:
            file_type = "original"
            try:
                with pdfplumber.open(upload_file) as pdf:
                    full_text = ""
                    for page in pdf.pages:
                        text = page.extract_text()
                        if text:
                            full_text += text + "\n"
            except Exception as e:
                st.warning(f"⚠️ Warning: Failed to open/parse PDF: {upload_file.name} -> {e}")
                failed_files.append(upload_file.name)
                continue

            if "This Purchase Order has been changed. Specific changes are shown in red." in full_text:
                file_type = "revised"
            po_info = parse_po_text(full_text, file_type=file_type)

            if not po_info:
                st.warning(
                    f"⚠️ Warning: Can not parse/extract information from this PDF file: {upload_file.name}, file type: {file_type}"
                )
                failed_files.append(upload_file.name)
                continue

            missing_keys = [k for k in required_keys if k not in po_info]
            if missing_keys:
                st.warning(
                    f"⚠️ Warning: PDF {upload_file.name} (file type: {file_type}) missing columns: {missing_keys}, skipped."
                )
                failed_files.append(upload_file.name)
                continue

            if file_type == "revised":
                revised_result_list.append(po_info)
                revised_files.append(upload_file.name)
            else:
                result_list.append(po_info)
                original_files.append(upload_file.name)

        if not (result_list or revised_result_list):
            st.error("Error: Can not successfully parse ANY PDF files, no report will be generated.")
            st.session_state.df = None
            st.session_state.excel_bytes = None
        else:
            original_df = pd.DataFrame(result_list)
            revised_df = pd.DataFrame(revised_result_list)
            if not revised_df.empty:
                original_df = pd.concat([original_df, revised_df], ignore_index=True)
            df = (
                original_df.drop_duplicates(subset=["PO#"], keep="last")
                .sort_values("PO#")
                .reset_index(drop=True)
            )

            st.session_state.info = {
                "original_files": original_files,
                "revised_files": revised_files,
                "failed_files": failed_files,
            }
            st.session_state.df = df
            st.session_state.excel_bytes = write_styled_excel(df)
            st.session_state.failed_files = failed_files

if st.session_state.df is not None:
    st.dataframe(st.session_state.df.head())

    st.download_button(
        label="📥 Download Excel",
        data=st.session_state.excel_bytes,
        file_name=f"new_orders_{datetime.now().date().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    if not st.session_state.info["failed_files"]:
        st.success("✅ All files parsed successfully!")
    else:
        st.info(
            f"Original files: {len(st.session_state.info['original_files'])}, revised files: {len(st.session_state.info['revised_files'])}"
        )
        st.warning(
            f"⚠️ Failed to parse some files below: {st.session_state.info['failed_files']}.\nPlease check them again."
        )

    st.success("Please download the Excel file.")
