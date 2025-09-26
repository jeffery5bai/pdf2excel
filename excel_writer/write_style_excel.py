import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

COL_LENGTH_OFFSET = 12


def write_styled_excel(df: pd.DataFrame):
    temp_file = "temp.xlsx"
    df.to_excel(temp_file, index=False, engine="openpyxl")

    wb = load_workbook(temp_file)
    ws = wb.active

    # Header style
    header_fill = PatternFill(
        start_color="F2DCDC", end_color="F2DCDC", fill_type="solid"
    )
    header_font = Font(name="Arial", size=12, bold=True)

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    # Data style
    data_font = Font(name="Arial", size=12)
    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.font = data_font

    # Set column format
    for row in range(2, ws.max_row + 1):
        # Data columns
        for col_name in ["Create Date", "Due Date", "GT CRD"]:
            col_idx = df.columns.get_loc(col_name) + 1
            cell = ws.cell(row=row, column=col_idx)
            if isinstance(cell.value, datetime):
                cell.number_format = "MM-DD-YYYY"

        # Number columns
        if "Qty" in df.columns:
            qty_idx = df.columns.get_loc("Qty") + 1
            cell = ws.cell(row=row, column=qty_idx)
            if isinstance(cell.value, str):
                cell.value = int(cell.value)
            cell.number_format = "0"

        if "Unit Price" in df.columns:
            unit_price_idx = df.columns.get_loc("Unit Price") + 1
            cell = ws.cell(row=row, column=unit_price_idx)
            if isinstance(cell.value, str):
                cell.value = float(cell.value)
            cell.number_format = "0.00"

    # adjust column width
    for col_idx, col_name in enumerate(df.columns, 1):
        max_length = (
            max(df[col_name].astype(str).map(len).max(), len(col_name))
            + COL_LENGTH_OFFSET
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length

    wb.save(temp_file)

    with open(temp_file, "rb") as f:
        return f.read()
